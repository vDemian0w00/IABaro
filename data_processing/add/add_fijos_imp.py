# Imports
import json
import openpyxl

# Funcion para agregar datos a fijos_imp.json

# Leer datos de excel
wb = openpyxl.load_workbook('../synthetic_data/data_synthetic_baro.xlsx')
hoja1 = wb['fijos_imp']

columna_names = []
for row in hoja1.iter_rows(values_only=True):
    columna_names.append(row[0])

columna_description = []
for row in hoja1.iter_rows(values_only=True):
    columna_description.append(row[1])
    
columna_amount = []
for row in hoja1.iter_rows(values_only=True):
    columna_amount.append(row[2])

# datos a agregar al archivo JSON
json_fijos_imp = '../classification/fijos_imp.json'
name_fijos_imp = columna_names
description_fijos_imp = columna_description
amount_fijos_imp = columna_amount

# leer los datos existentes del archivo JSON (si existen)
try:
    with open(json_fijos_imp, 'r', encoding='utf-8') as f:
        data_fijos_imp = json.load(f)
except FileNotFoundError:
    data_fijos_imp = []

# agregar nuevos datos al archivo JSON
for i, name in enumerate(name_fijos_imp):
    new_data = {
        'id': len(data_fijos_imp) + 1,
        'name': name,
        'description': description_fijos_imp[i],
        'amount': amount_fijos_imp[i],
    }
    data_fijos_imp.append(new_data)

# guardar los datos actualizados en el archivo JSON
with open(json_fijos_imp, 'w', encoding='utf-8') as f:
    json.dump(data_fijos_imp, f, indent=4)
