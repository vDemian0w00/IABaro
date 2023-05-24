# Imports
import json
import openpyxl

# Funcion para agregar datos a variables_imp.json

# Leer datos de excel
wb = openpyxl.load_workbook('../synthetic_data/data_synthetic_baro.xlsx')
hoja3 = wb['variables_imp']

columna_names = []
for row in hoja3.iter_rows(values_only=True):
    columna_names.append(row[0])

columna_description = []
for row in hoja3.iter_rows(values_only=True):
    columna_description.append(row[1])
    
columna_amount = []
for row in hoja3.iter_rows(values_only=True):
    columna_amount.append(row[2])

# datos a agregar al archivo JSON
json_variables_imp = '../classification/diarios_imp.json'
name_variables_imp = columna_names
description_variables_imp = columna_description
amount_variables_imp = columna_amount

# leer los datos existentes del archivo JSON (si existen)
try:
    with open(json_variables_imp, 'r', encoding='utf-8') as f:
        data_variables_imp = json.load(f)
except FileNotFoundError:
    data_variables_imp = []

# agregar nuevos datos al archivo JSON
for i, name in enumerate(name_variables_imp):
    new_data = {
        'id': len(data_variables_imp) + 1,
        'name': name,
        'description': description_variables_imp[i],
        'amount': amount_variables_imp[i],
    }
    data_variables_imp.append(new_data)

# guardar los datos actualizados en el archivo JSON
with open(json_variables_imp, 'w', encoding='utf-8') as f:
    json.dump(data_variables_imp, f, indent=4)
