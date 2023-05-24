# Imports
import json
import openpyxl

# Funcion para agregar datos a variables_pre.json

# Leer datos de excel
wb = openpyxl.load_workbook('../synthetic_data/data_synthetic_baro.xlsx')
hoja4 = wb['variables_pre']

columna_names = []
for row in hoja4.iter_rows(values_only=True):
    columna_names.append(row[0])

columna_description = []
for row in hoja4.iter_rows(values_only=True):
    columna_description.append(row[1])
    
columna_amount = []
for row in hoja4.iter_rows(values_only=True):
    columna_amount.append(row[2])

# datos a agregar al archivo JSON
json_variables_pre = '../classification/diarios_pre.json'
name_variables_pre = columna_names
description_variables_pre = columna_description
amount_variables_pre = columna_amount

# leer los datos existentes del archivo JSON (si existen)
try:
    with open(json_variables_pre, 'r', encoding='utf-8') as f:
        data_variables_pre = json.load(f)
except FileNotFoundError:
    data_variables_pre = []

# agregar nuevos datos al archivo JSON
for i, name in enumerate(name_variables_pre):
    new_data = {
        'id': len(data_variables_pre) + 1,
        'name': name,
        'description': description_variables_pre[i],
        'amount': amount_variables_pre[i],
    }
    data_variables_pre.append(new_data)

# guardar los datos actualizados en el archivo JSON
with open(json_variables_pre, 'w', encoding='utf-8') as f:
    json.dump(data_variables_pre, f, indent=4)