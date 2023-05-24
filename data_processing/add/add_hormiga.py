# Imports
import json
import openpyxl

# Funcion para agregar datos a hormiga.json

# Leer datos de excel
wb = openpyxl.load_workbook('../synthetic_data/data_synthetic_baro.xlsx')
hoja5 = wb['hormiga']

columna_names = []
for row in hoja5.iter_rows(values_only=True):
    columna_names.append(row[0])

columna_description = []
for row in hoja5.iter_rows(values_only=True):
    columna_description.append(row[1])
    
columna_amount = []
for row in hoja5.iter_rows(values_only=True):
    columna_amount.append(row[2])

# datos a agregar al archivo JSON
json_hormiga = '../classification/hormiga.json'
name_hormiga = columna_names
description_hormiga = columna_description
amount_hormiga = columna_amount

# leer los datos existentes del archivo JSON (si existen)
try:
    with open(json_hormiga, 'r', encoding='utf-8') as f:
        data_hormiga = json.load(f)
except FileNotFoundError:
    data_hormiga = []

# agregar nuevos datos al archivo JSON
for i, name in enumerate(name_hormiga):
    new_data = {
        'id': len(data_hormiga) + 1,
        'name': name,
        'description': description_hormiga[i],
        'amount': amount_hormiga[i],
    }
    data_hormiga.append(new_data)

# guardar los datos actualizados en el archivo JSON
with open(json_hormiga, 'w', encoding='utf-8') as f:
    json.dump(data_hormiga, f, indent=4)