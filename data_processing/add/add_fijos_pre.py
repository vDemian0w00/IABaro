# Imports
import json
import openpyxl

# Funcion para agregar datos a fijos_pre.json

# Leer datos de excel
wb = openpyxl.load_workbook('../synthetic_data/data_synthetic_baro.xlsx')
hoja2 = wb['fijos_pre']

columna_names = []
for row in hoja2.iter_rows(values_only=True):
    columna_names.append(row[0])

columna_description = []
for row in hoja2.iter_rows(values_only=True):
    columna_description.append(row[1])

columna_amount = []
for row in hoja2.iter_rows(values_only=True):
    columna_amount.append(row[2])

# datos a agregar al archivo JSON
json_fijos_pre = '../classification/fijos_pre.json'
name_fijos_pre = columna_names
description_fijos_pre = columna_description
amount_fijos_pre = columna_amount

# leer los datos existentes del archivo JSON (si existen)
try:
    with open(json_fijos_pre, 'r', encoding='utf-8') as f:
        data_fijos_pre = json.load(f)
except FileNotFoundError:
    data_fijos_pre = []

# agregar nuevos datos al archivo JSON
for i, name in enumerate(name_fijos_pre):
    new_data = {
        'id': len(data_fijos_pre) + 1,
        'name': name,
        'description': description_fijos_pre[i],
        'amount': amount_fijos_pre[i],
    }
    data_fijos_pre.append(new_data)

# guardar los datos actualizados en el archivo JSON
with open(json_fijos_pre, 'w', encoding='utf-8') as f:
    json.dump(data_fijos_pre, f, indent=4)
