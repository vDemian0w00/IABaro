# Imports
import json
import openpyxl

# Funcion para agregar datos a imprevistos.json

# Leer datos de excel
wb = openpyxl.load_workbook('../synthetic_data/data_synthetic_baro.xlsx')
hoja6 = wb['imprevistos']

columna_names = []
for row in hoja6.iter_rows(values_only=True):
    columna_names.append(row[0])

columna_description = []
for row in hoja6.iter_rows(values_only=True):
    columna_description.append(row[1])
    
columna_amount = []
for row in hoja6.iter_rows(values_only=True):
    columna_amount.append(row[2])

# datos a agregar al archivo JSON
json_imprevistos = '../classification/imprevistos.json'
name_imprevistos = columna_names
description_imprevistos = columna_description
amount_imprevistos = columna_amount

# leer los datos existentes del archivo JSON (si existen)
try:
    with open(json_imprevistos, 'r', encoding='utf-8') as f:
        data_imprevistos = json.load(f)
except FileNotFoundError:
    data_imprevistos = []

# agregar nuevos datos al archivo JSON
for i, name in enumerate(name_imprevistos):
    new_data = {
        'id': len(data_imprevistos) + 1,
        'name': name,
        'description': description_imprevistos[i],
        'amount': amount_imprevistos[i],
    }
    data_imprevistos.append(new_data)

# guardar los datos actualizados en el archivo JSON
with open(json_imprevistos, 'w', encoding='utf-8') as f:
    json.dump(data_imprevistos, f, indent=4)